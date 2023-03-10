from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
from datetime import date, timedelta,datetime

start_date = "2023-02-20"
end_date = "2023-02-25"

calls_dict = {
    "Della": {
        "total_calls": 0,
        "Monday": 0,
        "Tuesday": 0,
        "Wednesday": 0,
        "Thursday": 0,
        "Friday": 0
    },
    "Renee": {
        "total_calls": 0,
        "Monday": 0,
        "Tuesday": 0,
        "Wednesday": 0,
        "Thursday": 0,
        "Friday": 0
    },
    "Lyle": {
        "total_calls": 0,
        "Monday": 0,
        "Tuesday": 0,
        "Wednesday": 0,
        "Thursday": 0,
        "Friday": 0
    },
    "Janice": {
        "total_calls": 0,
        "Monday": 0,
        "Tuesday": 0,
        "Wednesday": 0,
        "Thursday": 0,
        "Friday": 0
    },
    "Foroza": {
        "total_calls": 0,
        "Monday": 0,
        "Tuesday": 0,
        "Wednesday": 0,
        "Thursday": 0,
        "Friday": 0
    },
    "Denise": {
        "total_calls": 0,
        "Monday": 0,
        "Tuesday": 0,
        "Wednesday": 0,
        "Thursday": 0,
        "Friday": 0
    }
}


def dates(start,end):
    start_date_box = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[33]/div[2]/div[1]/div/div/div/div/div/div/div/div/div/div[2]/div[2]/div/div[3]/div/div/div[3]/div/div[1]/div[2]/div/form/div/div/div/div[1]/div[2]/div/fieldset/div/div/fieldset[1]/div/div/div[1]/div[1]/div/div/div[1]/input')))
    start_date_box.clear()
    start_date_box.send_keys(str(start))
    end_date_box = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[33]/div[2]/div[1]/div/div/div/div/div/div/div/div/div/div[2]/div[2]/div/div[3]/div/div/div[3]/div/div[1]/div[2]/div/form/div/div/div/div[1]/div[2]/div/fieldset/div/div/fieldset[1]/div/div/div[2]/div[1]/div/div/div[1]/input')))
    end_date_box.clear()
    end_date_box.send_keys(str(end))

#list below hidden to protect client
extention_list = .......

# set the path to the chromedriver executable
chromedriver_path = "/Users/bladebrink/Library/Mobile Documents/com~apple~CloudDocs/SchoolofIT/chromedriver"

# create a Chrome webdriver instance
driver = webdriver.Chrome(chromedriver_path)

# navigate to the website
driver.get("https://web.vodcs.co.za:8444/")

def pulling():
    show_transactions = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH,"/html/body/div[33]/div[2]/div[1]/div/div/div/div/div/div/div/div/div/div[2]/div[2]/div/div[3]/div/div/div[3]/div/div[1]/div[2]/div/form/div/div/div/div[1]/div[2]/div/fieldset/div/div/fieldset[3]/div/div/div/table[1]/tbody/tr[2]/td[2]/em/button")))
    show_transactions.click()
    show_transactions.click()
    show_transactions.click()
    show_transactions.click()
    time.sleep(5)
    try:
        count = driver.find_element_by_xpath("/html/body/div[33]/div[2]/div[1]/div/div/div/div/div/div/div/div/div/div[2]/div[2]/div/div[3]/div/div/div[3]/div/div[1]/div[2]/div/form/div/div/div/div[2]/div/div/fieldset/div/div/div/div[2]/div[1]/div/div[1]/div[2]/div/div/table/tbody/tr/td[2]/div")
        # Get the text value of the element
        value = count.text
        total = value
    except:
        total = 0
    return total

# wait for the username box element to be located for up to 5 seconds
username_box = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/div[2]/div[1]/div/div/div/div/div/form/div[1]/div[1]/input')))

username_box.send_keys("User Name")
time.sleep(20) #waiting for password input

for y in range(len(extention_list)):
    if y>0:
        login_button = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/div[2]/div[1]/div/div/div/div/div/form/table/tbody/tr[2]/td[2]')))
        login_button.click()
        time.sleep(10) #waiting for home page to load
    if y == 0:
        person = "Della"
    elif y == 1:
        person = "Renee"
    elif y == 2:
        person = "Lyle"
    elif y == 3:
        person = "Janice"
    elif y == 4:
        person = "Foroza"
    elif y == 5:
        person = "Denise"

    ip_centrix = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'/html/body/div[2]/div/div/div[2]/div[1]/div[3]/ul/li[3]/a[2]/em/span/span')))
    ip_centrix.click()
    
    phone_lines = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[2]/div/div/div[2]/div[2]/div/div[3]/div/div/div[2]/div[2]/div/ul/div/li[3]/div')))
    phone_lines.click()
    
    extention_box = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div/div/div[2]/div[2]/div/div[3]/div/div/div[3]/div/div[1]/div[2]/div/form/div/div/div[1]/div/table/tbody/tr/td[1]/table/tbody/tr/td[2]/input')))
    extention_box.send_keys(extention_list[y])
    # Wait for the overlay to disappear
    # WebDriverWait(driver, 20).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".ext-el-mask")))
    time.sleep(4)
    # Click the element
    search_btn = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div/div/div[2]/div[2]/div/div[3]/div/div/div[3]/div/div[1]/div[2]/div/form/div/div/div[1]/div/table/tbody/tr/td[2]/table/tbody/tr/td[1]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/em/button')))
    search_btn.click()
    time.sleep(7)
    access_btn = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div/div/div[2]/div[2]/div/div[3]/div/div/div[3]/div/div[1]/div[2]/div/form/div/div/div[2]/div/div[1]/div[2]/div/div/table/tbody/tr/td[1]/div/a/img')))
    time.sleep(3)
    access_btn.click()
    time.sleep(3)
    billing_info = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[33]/div[2]/div[1]/div/div/div/div/div/div/div/div/div/div[2]/div[1]/div[3]/ul/li[5]/a[2]/em/span/span')))
    billing_info.click()
    transactions = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[33]/div[2]/div[1]/div/div/div/div/div/div/div/div/div/div[2]/div[2]/div/div[3]/div/div/div[2]/div[2]/div/ul/div/li[7]/div/a/span')))
    transactions.click()
    time.sleep(3)
    end_time_box = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[33]/div[2]/div[1]/div/div/div/div/div/div/div/div/div/div[2]/div[2]/div/div[3]/div/div/div[3]/div/div[1]/div[2]/div/form/div/div/div/div[1]/div[2]/div/fieldset/div/div/fieldset[1]/div/div/div[2]/div[1]/div/div/div[3]/input')))
    end_time_box.clear()
    end_time_box.send_keys("13:00:00")
    end_time_box.send_keys(Keys.RETURN)
    start_time_box = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[33]/div[2]/div[1]/div/div/div/div/div/div/div/div/div/div[2]/div[2]/div/div[3]/div/div/div[3]/div/div[1]/div[2]/div/form/div/div/div/div[1]/div[2]/div/fieldset/div/div/fieldset[1]/div/div/div[1]/div[1]/div/div/div[3]/input')))
    start_time_box.clear()
    start_time_box.send_keys("00:00:00")
    start_time_box.send_keys(Keys.RETURN)

    checkbox = driver.find_element_by_xpath("/html/body/div[33]/div[2]/div[1]/div/div/div/div/div/div/div/div/div/div[2]/div[2]/div/div[3]/div/div/div[3]/div/div[1]/div[2]/div/form/div/div/div/div[1]/div[2]/div/fieldset/div/div/fieldset[2]/div/div/div[3]/div[1]/div/input")
    checkbox.click()

    #the first date entry
    dates(start_date,end_date)

    #the first pull
    value = pulling()
    calls_dict[person]["total_calls"] = value
   
    
     #the second date entry
    dates(start_date,start_date)
    

    #the second pull
    value = pulling()
    calls_dict[person]["Monday"] = value
    
    time.sleep(5)
    time.sleep(2)


    date1 = datetime.strptime(start_date, "%Y-%m-%d")
    for x in range(4):  
        if x == 0:
            day = "Tuesday"
        elif x == 1:
            day = "Wednesday"
        elif x == 2:
            day = "Thursday"
        elif x == 3:
            day = "Friday"
       

        # Input the number of days you want to increase the date by
        num_days = 1

        # Add the number of days to the date
        new_date = date1 + timedelta(days=num_days)

        date1 = new_date 
        #the final dates entries
        dates(new_date.strftime("%Y-%m-%d"),new_date.strftime("%Y-%m-%d"))
        #the final pulls
        value = pulling()
    
        calls_dict[person][day] = value
    

    driver.refresh()
    time.sleep(5)
   
    

wb = openpyxl.load_workbook('Canvasser Report.xlsx')

# Select the worksheet
ws = wb.active

# Edit certain cells
ws["B1"] = start_date
ws["B2"] = end_date
ws['C9'] = int(calls_dict['Della']['total_calls'])
ws['G9'] = int(calls_dict['Della']['Monday'])
ws['H9'] = int(calls_dict['Della']['Tuesday'])
ws['I9'] = int(calls_dict['Della']['Wednesday'])
ws['J9'] = int(calls_dict['Della']['Thursday'])
ws['K9'] = int(calls_dict['Della']['Friday'])

ws['C8'] = int(calls_dict['Denise']['total_calls'])
ws['G8'] = int(calls_dict['Denise']['Monday'])
ws['H8'] = int(calls_dict['Denise']['Tuesday'])
ws['I8'] = int(calls_dict['Denise']['Wednesday'])
ws['J8'] = int(calls_dict['Denise']['Thursday'])
ws['K8'] = int(calls_dict['Denise']['Friday'])

ws['C7'] = int(calls_dict['Lyle']['total_calls'])
ws['G7'] = int(calls_dict['Lyle']['Monday'])
ws['H7'] = int(calls_dict['Lyle']['Tuesday'])
ws['I7'] = int(calls_dict['Lyle']['Wednesday'])
ws['J7'] = int(calls_dict['Lyle']['Thursday'])
ws['K7'] = int(calls_dict['Lyle']['Friday'])

ws['C6'] = int(calls_dict['Janice']['total_calls'])
ws['G6'] = int(calls_dict['Janice']['Monday'])
ws['H6'] = int(calls_dict['Janice']['Tuesday'])
ws['I6'] = int(calls_dict['Janice']['Wednesday'])
ws['J6'] = int(calls_dict['Janice']['Thursday'])
ws['K6'] = int(calls_dict['Janice']['Friday'])

ws['C5'] = int(calls_dict['Foroza']['total_calls'])
ws['G5'] = int(calls_dict['Foroza']['Monday'])
ws['H5'] = int(calls_dict['Foroza']['Tuesday'])
ws['I5'] = int(calls_dict['Foroza']['Wednesday'])
ws['J5'] = int(calls_dict['Foroza']['Thursday'])
ws['K5'] = int(calls_dict['Foroza']['Friday'])

ws['C4'] = int(calls_dict['Renee']['total_calls'])
ws['G4'] = int(calls_dict['Renee']['Monday'])
ws['H4'] = int(calls_dict['Renee']['Tuesday'])
ws['I4'] = int(calls_dict['Renee']['Wednesday'])
ws['J4'] = int(calls_dict['Renee']['Thursday'])
ws['K4'] = int(calls_dict['Renee']['Friday'])

filename = "Canvasser Report.xlsx"
# Save the changes
wb.save(filename)

# Close the Excel file
wb.close()
time.sleep(5)
driver.quit()


